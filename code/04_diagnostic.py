import json
import os
import sys
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.image import Image as XLImage

PROJECT_ROOT = Path(r"C:\Users\Ho On Tam\Desktop\bot\weather")
MODEL_ROOT_DIR = PROJECT_ROOT / "data" / "Model"
SCRIPT_NAME = "04_diagnostic.py"


def pick_run_dir(model_root: Path, run_date: str | None = None) -> Path:
    if run_date:
        candidate = model_root / run_date
        if not candidate.exists():
            raise FileNotFoundError(f"Requested run folder does not exist: {candidate}")
        return candidate

    env_run_date = os.getenv("MODEL_RUN_DATE")
    if env_run_date:
        candidate = model_root / env_run_date
        if not candidate.exists():
            raise FileNotFoundError(f"MODEL_RUN_DATE points to missing folder: {candidate}")
        return candidate

    if len(sys.argv) > 1:
        candidate = model_root / sys.argv[1]
        if not candidate.exists():
            raise FileNotFoundError(f"Run folder from CLI arg does not exist: {candidate}")
        return candidate

    today_candidate = model_root / datetime.now().strftime("%Y%m%d")
    if (today_candidate / "03_test_predictions.parquet").exists():
        return today_candidate

    dated_dirs = sorted(
        [p for p in model_root.iterdir() if p.is_dir() and p.name.isdigit()],
        key=lambda p: p.name,
        reverse=True,
    )
    for p in dated_dirs:
        if (p / "03_test_predictions.parquet").exists():
            return p

    if dated_dirs:
        return dated_dirs[0]
    raise FileNotFoundError(f"No dated model folders found under {model_root}")


RUN_OUTPUT_DIR = pick_run_dir(MODEL_ROOT_DIR)
TEST_PREDICTIONS_PATH = RUN_OUTPUT_DIR / "03_test_predictions.parquet"
PROCESSED_PATH = RUN_OUTPUT_DIR / "02_processed_dataset.parquet"
METRICS_PATH = RUN_OUTPUT_DIR / "03_metrics.csv"
METRICS_BY_CITY_PATH = RUN_OUTPUT_DIR / "03_metrics_by_city.csv"
METRICS_BY_MONTH_PATH = RUN_OUTPUT_DIR / "03_metrics_by_month.csv"
METRICS_BY_HOUR_PATH = RUN_OUTPUT_DIR / "03_metrics_by_hour.csv"
PERM_IMPORTANCE_PATH = RUN_OUTPUT_DIR / "03_permutation_importance.csv"
SPLIT_SUMMARY_PATH = RUN_OUTPUT_DIR / "03_split_summary.json"
MODEL_SUMMARY_PATH = RUN_OUTPUT_DIR / "03_model_summary.json"

DIAG_DIR = RUN_OUTPUT_DIR / "Diagnostics"
SUMMARY_JSON = DIAG_DIR / "04_diagnostics_summary.json"
WORKBOOK_PATH = DIAG_DIR / "04_diagnostics_workbook.xlsx"

HORIZON_ORDER = [f"h{i}" for i in range(9, 0, -1)]


def ensure_dirs() -> None:
    DIAG_DIR.mkdir(parents=True, exist_ok=True)



def load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)



def _infer_group_col(df: pd.DataFrame) -> str:
    for col in ["state", "city", "airport", "location"]:
        if col in df.columns:
            return col
    return "city"



def load_data() -> tuple[pd.DataFrame, pd.DataFrame | None, str]:
    if not TEST_PREDICTIONS_PATH.exists():
        available = sorted([p.name for p in MODEL_ROOT_DIR.iterdir() if p.is_dir()]) if MODEL_ROOT_DIR.exists() else []
        raise FileNotFoundError(
            "Missing test predictions: "
            f"{TEST_PREDICTIONS_PATH}\n"
            f"Detected run folder: {RUN_OUTPUT_DIR}\n"
            f"Available dated folders: {available}\n"
            "Set MODEL_RUN_DATE=YYYYMMDD or run: python 04_diagnostic.py YYYYMMDD"
        )

    test_pred = pd.read_parquet(TEST_PREDICTIONS_PATH)
    if "timestamp" not in test_pred.columns:
        raise ValueError("03_test_predictions.parquet must contain timestamp.")
    test_pred["timestamp"] = pd.to_datetime(test_pred["timestamp"], errors="coerce")
    test_pred = test_pred[test_pred["timestamp"].notna()].copy()

    processed = None
    if PROCESSED_PATH.exists():
        processed = pd.read_parquet(PROCESSED_PATH)
        if "timestamp" in processed.columns:
            processed["timestamp"] = pd.to_datetime(processed["timestamp"], errors="coerce")
            processed = processed[processed["timestamp"].notna()].copy()

    group_col = _infer_group_col(test_pred)
    return test_pred, processed, group_col



def add_time_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["date"] = out["timestamp"].dt.date
    out["hour_num"] = out.get("hour_num", out["timestamp"].dt.hour)
    out["month_num"] = out.get("month_num", out["timestamp"].dt.month)
    return out



def _derive_horizon_bucket(hours_from_peak: pd.Series) -> pd.Series:
    vals = pd.to_numeric(hours_from_peak, errors="coerce")
    out = pd.Series(pd.NA, index=vals.index, dtype="object")
    for h in range(1, 10):
        lo = -h
        hi = -(h - 1)
        mask = (vals >= lo) & (vals < hi)
        out.loc[mask] = f"h{h}"
    return pd.Categorical(out, categories=HORIZON_ORDER, ordered=True)



def add_error_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "pred_daily_max_mean" not in out.columns and {"temp", "pred_remaining_heat_mean"}.issubset(out.columns):
        out["pred_daily_max_mean"] = out["temp"] + out["pred_remaining_heat_mean"]
    for q in ["q50", "q90", "q95"]:
        pred_col = f"pred_remaining_heat_{q}"
        daily_col = f"pred_daily_max_{q}"
        if daily_col not in out.columns and {"temp", pred_col}.issubset(out.columns):
            out[daily_col] = out["temp"] + out[pred_col]

    if "remaining_heat" in out.columns and "pred_remaining_heat_mean" in out.columns:
        out["remaining_heat_error"] = out["pred_remaining_heat_mean"] - out["remaining_heat"]
        out["remaining_heat_abs_error"] = out["remaining_heat_error"].abs()

    if "daily_max_temp" in out.columns and "pred_daily_max_mean" in out.columns:
        out["daily_max_error"] = out["pred_daily_max_mean"] - out["daily_max_temp"]
        out["daily_max_abs_error"] = out["daily_max_error"].abs()

    if "hours_from_solar_peak" in out.columns and "horizon_bucket" not in out.columns:
        out["horizon_bucket"] = _derive_horizon_bucket(out["hours_from_solar_peak"])
    elif "hours_to_peak" in out.columns and "horizon_bucket" not in out.columns:
        out["horizon_bucket"] = pd.cut(
            pd.to_numeric(out["hours_to_peak"], errors="coerce"),
            bins=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, np.inf],
            labels=["h1", "h2", "h3", "h4", "h5", "h6", "h7", "h8", "h9", "h9plus"],
            right=False,
        )
    if "horizon_bucket" in out.columns:
        out["horizon_bucket"] = pd.Categorical(out["horizon_bucket"].astype(str), categories=HORIZON_ORDER + ["h9plus"], ordered=True)

    return out



def summarize_error(df: pd.DataFrame, group_cols: list[str], error_col: str) -> pd.DataFrame:
    if error_col not in df.columns:
        cols = group_cols + ["rows", "mae", "rmse", "bias"]
        return pd.DataFrame(columns=cols)
    work = df.dropna(subset=[error_col]).copy()
    if work.empty:
        cols = group_cols + ["rows", "mae", "rmse", "bias"]
        return pd.DataFrame(columns=cols)
    grp = work.groupby(group_cols, dropna=False, observed=False)
    out = grp[error_col].agg(
        rows="size",
        mae=lambda x: float(np.mean(np.abs(x))),
        rmse=lambda x: float(np.sqrt(np.mean(np.square(x)))),
        bias="mean",
    ).reset_index()
    return out



def peak_capture_curve(df: pd.DataFrame) -> pd.DataFrame:
    if "daily_max_error" not in df.columns or "horizon_bucket" not in df.columns:
        return pd.DataFrame(columns=["horizon_bucket", "within_0p5", "within_1p0", "rows"])
    work = df.dropna(subset=["daily_max_error", "horizon_bucket"]).copy()
    if work.empty:
        return pd.DataFrame(columns=["horizon_bucket", "within_0p5", "within_1p0", "rows"])
    rows = []
    for hb, g in work.groupby("horizon_bucket", observed=False):
        if pd.isna(hb) or len(g) == 0:
            continue
        err = g["daily_max_error"].abs()
        rows.append(
            {
                "horizon_bucket": str(hb),
                "rows": int(len(g)),
                "within_0p5": float((err <= 0.5).mean()),
                "within_1p0": float((err <= 1.0).mean()),
            }
        )
    out = pd.DataFrame(rows)
    if not out.empty:
        out["horizon_bucket"] = pd.Categorical(out["horizon_bucket"], categories=HORIZON_ORDER, ordered=True)
        out = out.sort_values("horizon_bucket").assign(horizon_bucket=lambda x: x["horizon_bucket"].astype(str))
    return out



def quantile_calibration(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if "remaining_heat" not in df.columns:
        return pd.DataFrame(columns=["quantile_label", "nominal_quantile", "observed_coverage", "coverage_gap"])
    for label, q in [("q50", 0.50), ("q90", 0.90), ("q95", 0.95)]:
        col = f"pred_remaining_heat_{label}"
        if col not in df.columns:
            continue
        coverage = float((df["remaining_heat"] <= df[col]).mean())
        rows.append(
            {
                "quantile_label": label,
                "nominal_quantile": q,
                "observed_coverage": coverage,
                "coverage_gap": coverage - q,
            }
        )
    return pd.DataFrame(rows)



def save_plot(path: Path) -> None:
    plt.tight_layout()
    plt.savefig(path, dpi=160, bbox_inches="tight")
    plt.close()



def plot_horizon_curve(df_plot: pd.DataFrame, group_col: str, value_col: str, title: str, ylabel: str, out_name: str) -> None:
    if df_plot.empty:
        return
    plt.figure(figsize=(10, 6))
    for name, g in df_plot.groupby(group_col, observed=False):
        g2 = g.copy()
        g2["horizon_bucket"] = pd.Categorical(g2["horizon_bucket"], categories=HORIZON_ORDER, ordered=True)
        g2 = g2.sort_values("horizon_bucket")
        plt.plot(g2["horizon_bucket"].astype(str), g2[value_col], marker="o", label=str(name))
    plt.xlabel("Horizon bucket")
    plt.ylabel(ylabel)
    plt.title(title)
    if df_plot[group_col].nunique() <= 15:
        plt.legend(fontsize=8)
    save_plot(DIAG_DIR / out_name)



def plot_box_by_horizon(df: pd.DataFrame, error_col: str, out_name: str, title: str) -> None:
    if error_col not in df.columns or "horizon_bucket" not in df.columns:
        return
    data = []
    labels = []
    for h in HORIZON_ORDER:
        vals = df.loc[df["horizon_bucket"].astype(str) == h, error_col].dropna()
        if len(vals):
            data.append(vals.values)
            labels.append(h)
    if not data:
        return
    plt.figure(figsize=(10, 6))
    plt.boxplot(data, tick_labels=labels, showfliers=False)
    plt.axhline(0.0)
    plt.xlabel("Horizon bucket")
    plt.ylabel("Predicted max - true max")
    plt.title(title)
    save_plot(DIAG_DIR / out_name)



def plot_daily_max_scatter(df: pd.DataFrame, out_name: str) -> None:
    req = {"pred_daily_max_mean", "daily_max_temp"}
    if not req.issubset(df.columns):
        return
    sample = df.dropna(subset=list(req)).copy()
    if sample.empty:
        return
    if len(sample) > 20000:
        sample = sample.sample(20000, random_state=42)
    plt.figure(figsize=(7, 6))
    plt.scatter(sample["pred_daily_max_mean"], sample["daily_max_temp"], s=8, alpha=0.20)
    lo = min(sample["pred_daily_max_mean"].min(), sample["daily_max_temp"].min())
    hi = max(sample["pred_daily_max_mean"].max(), sample["daily_max_temp"].max())
    plt.plot([lo, hi], [lo, hi])
    plt.xlabel("Predicted daily max")
    plt.ylabel("True daily max")
    plt.title("Predicted vs True Daily Max")
    save_plot(DIAG_DIR / out_name)



def plot_heatmap(pivot_df: pd.DataFrame, title: str, out_name: str, cbar_label: str) -> None:
    if pivot_df.empty:
        return
    plt.figure(figsize=(10, max(4, 0.4 * len(pivot_df) + 2)))
    arr = pivot_df.to_numpy(dtype=float)
    im = plt.imshow(arr, aspect="auto")
    plt.xticks(range(len(pivot_df.columns)), pivot_df.columns, rotation=45)
    plt.yticks(range(len(pivot_df.index)), pivot_df.index)
    plt.title(title)
    cbar = plt.colorbar(im)
    cbar.set_label(cbar_label)
    save_plot(DIAG_DIR / out_name)



def plot_hour_curve(df_plot: pd.DataFrame, value_col: str, title: str, ylabel: str, out_name: str) -> None:
    if df_plot.empty:
        return
    plt.figure(figsize=(10, 5))
    plt.plot(df_plot["hour_num"], df_plot[value_col], marker="o")
    if value_col == "bias":
        plt.axhline(0.0)
    plt.xlabel("Hour of day")
    plt.ylabel(ylabel)
    plt.title(title)
    save_plot(DIAG_DIR / out_name)



def plot_remaining_heat_scatter(df: pd.DataFrame, out_name: str) -> None:
    req = {"pred_remaining_heat_mean", "remaining_heat"}
    if not req.issubset(df.columns):
        return
    sample = df.dropna(subset=list(req)).copy()
    if sample.empty:
        return
    if len(sample) > 20000:
        sample = sample.sample(20000, random_state=42)
    plt.figure(figsize=(7, 6))
    plt.scatter(sample["pred_remaining_heat_mean"], sample["remaining_heat"], s=8, alpha=0.20)
    lo = min(sample["pred_remaining_heat_mean"].min(), sample["remaining_heat"].min())
    hi = max(sample["pred_remaining_heat_mean"].max(), sample["remaining_heat"].max())
    plt.plot([lo, hi], [lo, hi])
    plt.xlabel("Predicted remaining heat")
    plt.ylabel("True remaining heat")
    plt.title("Remaining Heat Calibration")
    save_plot(DIAG_DIR / out_name)



def plot_peak_capture(df_plot: pd.DataFrame, out_name: str) -> None:
    if df_plot.empty:
        return
    plt.figure(figsize=(10, 5))
    plt.plot(df_plot["horizon_bucket"], df_plot["within_0p5"], marker="o", label="Within ±0.5°C")
    plt.plot(df_plot["horizon_bucket"], df_plot["within_1p0"], marker="o", label="Within ±1.0°C")
    plt.xlabel("Horizon bucket")
    plt.ylabel("Share captured")
    plt.title("Peak Capture Curve")
    plt.legend()
    save_plot(DIAG_DIR / out_name)



def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, len(val))
        ws.column_dimensions[col].width = min(length + 2, 40)



def _excel_safe_value(val):
    if isinstance(val, np.generic):
        val = val.item()
    if isinstance(val, (list, tuple, dict, set, np.ndarray)):
        try:
            return json.dumps(val if not isinstance(val, np.ndarray) else val.tolist(), ensure_ascii=False)
        except Exception:
            return str(val)
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    return val


def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, index: bool = False) -> tuple[int, int]:
    frame = df.copy()
    if index:
        frame = frame.reset_index()
    for j, col in enumerate(frame.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    for i, (_, row) in enumerate(frame.iterrows(), start=start_row + 1):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=_excel_safe_value(val))
    autosize_columns(ws)
    return start_row + 1, start_col + len(frame.columns) - 1



def add_line_chart(ws, title: str, y_title: str, x_title: str, min_row: int, max_row: int, min_col: int, max_col: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.width = 14
    chart.height = 8
    chart.y_axis.majorGridlines = ChartLines()
    data = Reference(ws, min_col=min_col + 1, max_col=max_col, min_row=min_row - 1, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)



def add_scatter_chart(ws, title: str, x_title: str, y_title: str, x_col: int, y_col: int, min_row: int, max_row: int, anchor: str) -> None:
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.width = 12
    chart.height = 8
    xvalues = Reference(ws, min_col=x_col, min_row=min_row, max_row=max_row)
    yvalues = Reference(ws, min_col=y_col, min_row=min_row, max_row=max_row)
    series = Series(yvalues, xvalues, title_from_data=False, title="points")
    chart.series.append(series)
    ws.add_chart(chart, anchor)



def build_workbook(tables: dict[str, pd.DataFrame], images: list[Path]) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet_name, df in tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        r0, cmax = write_df(ws, df)
        n_rows = len(df) + 1
        if sheet_name == "Horizon_MAE" and not df.empty:
            add_line_chart(ws, "Horizon Accuracy Curve", "MAE", "Horizon", 2, n_rows, 1, cmax, "H2")
        elif sheet_name == "Horizon_Bias" and not df.empty:
            add_line_chart(ws, "Bias by Horizon", "Bias", "Horizon", 2, n_rows, 1, cmax, "H2")
        elif sheet_name == "Peak_Capture" and not df.empty:
            add_line_chart(ws, "Peak Capture Curve", "Share", "Horizon", 2, n_rows, 1, cmax, "F2")
        elif sheet_name == "Pred_vs_Actual" and not df.empty:
            x_col = list(df.columns).index("pred_daily_max_mean") + 1
            y_col = list(df.columns).index("daily_max_temp") + 1
            add_scatter_chart(ws, "Predicted vs True Daily Max", "Predicted", "True", x_col, y_col, 2, n_rows, "F2")

    img_ws = wb.create_sheet("Chart_Images")
    row_cursor = 1
    for img_path in images:
        if img_path.exists():
            img_ws.cell(row=row_cursor, column=1, value=img_path.name)
            img = XLImage(str(img_path))
            img.width = 900
            img.height = 500
            img_ws.add_image(img, f"A{row_cursor + 1}")
            row_cursor += 28
    wb.save(WORKBOOK_PATH)



def main() -> None:
    ensure_dirs()
    print(f"Using run folder: {RUN_OUTPUT_DIR}")

    df, processed, group_col = load_data()
    df = add_time_columns(df)
    df = add_error_columns(df)

    if "daily_max_error" not in df.columns:
        raise ValueError(
            "Need daily max diagnostics columns. Could not find/create 'daily_max_error'. "
            "Expected at least: daily_max_temp, temp, pred_remaining_heat_mean."
        )
    if "horizon_bucket" not in df.columns:
        raise ValueError(
            "Could not derive horizon_bucket. Expected one of: horizon_bucket, hours_from_solar_peak, hours_to_peak."
        )

    horizon_stats = summarize_error(df, ["horizon_bucket", group_col], "daily_max_error")
    horizon_mae_wide = horizon_stats.pivot(index="horizon_bucket", columns=group_col, values="mae").reindex(HORIZON_ORDER)
    horizon_bias_wide = horizon_stats.pivot(index="horizon_bucket", columns=group_col, values="bias").reindex(HORIZON_ORDER)

    city_horizon_mae = summarize_error(df, [group_col, "horizon_bucket"], "daily_max_error")
    city_horizon_mae_pivot = city_horizon_mae.pivot(index=group_col, columns="horizon_bucket", values="mae").reindex(columns=HORIZON_ORDER)
    city_horizon_bias_pivot = city_horizon_mae.pivot(index=group_col, columns="horizon_bucket", values="bias").reindex(columns=HORIZON_ORDER)

    hour_stats = summarize_error(df, ["hour_num"], "daily_max_error").sort_values("hour_num")
    peak_capture = peak_capture_curve(df)
    qcal = quantile_calibration(df)

    pred_vs_actual = df[["pred_daily_max_mean", "daily_max_temp", "horizon_bucket", group_col]].dropna().copy()
    if len(pred_vs_actual) > 20000:
        pred_vs_actual = pred_vs_actual.sample(20000, random_state=42)

    remaining_heat_sample = pd.DataFrame()
    if {"pred_remaining_heat_mean", "remaining_heat", "horizon_bucket", group_col}.issubset(df.columns):
        remaining_heat_sample = df[["pred_remaining_heat_mean", "remaining_heat", "horizon_bucket", group_col]].dropna().copy()
        if len(remaining_heat_sample) > 20000:
            remaining_heat_sample = remaining_heat_sample.sample(20000, random_state=42)

    out_csvs = {
        "04_horizon_mae_by_group.csv": horizon_mae_wide.reset_index(),
        "04_horizon_bias_by_group.csv": horizon_bias_wide.reset_index(),
        "04_group_horizon_mae_heatmap.csv": city_horizon_mae_pivot.reset_index(),
        "04_group_horizon_bias_heatmap.csv": city_horizon_bias_pivot.reset_index(),
        "04_error_by_hour_daily_max.csv": hour_stats,
        "04_peak_capture_curve.csv": peak_capture,
        "04_quantile_calibration.csv": qcal,
        "04_pred_vs_actual_daily_max_sample.csv": pred_vs_actual,
        "04_remaining_heat_calibration_sample.csv": remaining_heat_sample,
    }
    for name, data in out_csvs.items():
        data.to_csv(DIAG_DIR / name, index=False)

    plot_horizon_curve(
        horizon_stats,
        group_col=group_col,
        value_col="mae",
        title=f"Horizon Accuracy Curve by {group_col.title()}",
        ylabel="MAE of predicted daily max",
        out_name="04_horizon_accuracy_curve.png",
    )
    plot_horizon_curve(
        horizon_stats,
        group_col=group_col,
        value_col="bias",
        title=f"Bias by Horizon by {group_col.title()}",
        ylabel="Mean error of predicted daily max",
        out_name="04_horizon_bias_curve.png",
    )
    plot_box_by_horizon(df, "daily_max_error", "04_daily_max_error_boxplot.png", "Distribution of Daily Max Error by Horizon")
    plot_daily_max_scatter(df, "04_predicted_vs_actual_daily_max.png")
    plot_heatmap(city_horizon_mae_pivot, f"{group_col.title()} Difficulty Heatmap (MAE)", "04_group_horizon_mae_heatmap.png", "MAE")
    plot_heatmap(city_horizon_bias_pivot, f"{group_col.title()} Bias Heatmap", "04_group_horizon_bias_heatmap.png", "Bias")
    plot_hour_curve(hour_stats, "mae", "Error vs Time of Day (Daily Max MAE)", "MAE", "04_error_vs_hour_mae.png")
    plot_hour_curve(hour_stats, "bias", "Bias vs Time of Day (Daily Max)", "Bias", "04_error_vs_hour_bias.png")
    plot_remaining_heat_scatter(df, "04_remaining_heat_calibration.png")
    plot_peak_capture(peak_capture, "04_peak_capture_curve.png")

    summary = {
        "run_output_dir": str(RUN_OUTPUT_DIR),
        "group_col": group_col,
        "n_test_rows": int(len(df)),
        "horizons_present": sorted([str(x) for x in df["horizon_bucket"].dropna().astype(str).unique()]),
        "test_start": str(df["timestamp"].min()),
        "test_end": str(df["timestamp"].max()),
        "overall_daily_max_mae": float(df["daily_max_abs_error"].mean()),
        "overall_daily_max_rmse": float(np.sqrt(np.mean(np.square(df["daily_max_error"])))),
        "overall_daily_max_bias": float(df["daily_max_error"].mean()),
        "split_summary": load_json(SPLIT_SUMMARY_PATH),
        "model_summary": load_json(MODEL_SUMMARY_PATH),
    }
    with open(SUMMARY_JSON, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    tables = {
        "Horizon_MAE": horizon_mae_wide.reset_index().rename(columns={"horizon_bucket": "horizon"}),
        "Horizon_Bias": horizon_bias_wide.reset_index().rename(columns={"horizon_bucket": "horizon"}),
        f"{group_col.title()}_MAE_Heatmap": city_horizon_mae_pivot.reset_index(),
        f"{group_col.title()}_Bias_Heatmap": city_horizon_bias_pivot.reset_index(),
        "Error_By_Hour": hour_stats,
        "Peak_Capture": peak_capture,
        "Quantile_Calibration": qcal,
        "Pred_vs_Actual": pred_vs_actual,
        "Remaining_Heat_Calib": remaining_heat_sample,
        "Summary": pd.DataFrame([summary]),
    }
    image_files = [
        DIAG_DIR / "04_horizon_accuracy_curve.png",
        DIAG_DIR / "04_horizon_bias_curve.png",
        DIAG_DIR / "04_daily_max_error_boxplot.png",
        DIAG_DIR / "04_predicted_vs_actual_daily_max.png",
        DIAG_DIR / "04_group_horizon_mae_heatmap.png",
        DIAG_DIR / "04_group_horizon_bias_heatmap.png",
        DIAG_DIR / "04_error_vs_hour_mae.png",
        DIAG_DIR / "04_error_vs_hour_bias.png",
        DIAG_DIR / "04_remaining_heat_calibration.png",
        DIAG_DIR / "04_peak_capture_curve.png",
    ]
    build_workbook(tables, image_files)

    print(f"Saved diagnostics folder -> {DIAG_DIR}")
    print(f"Saved workbook -> {WORKBOOK_PATH}")
    print(f"Saved summary -> {SUMMARY_JSON}")
    print("04_diagnostic complete.")


if __name__ == "__main__":
    main()
